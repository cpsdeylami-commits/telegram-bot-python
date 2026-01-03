from telegram import Update, ReplyKeyboardMarkup
from telegram.ext import ApplicationBuilder, CommandHandler, ContextTypes
from telegram import InlineKeyboardMarkup, InlineKeyboardButton
from telegram import InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import CallbackQueryHandler, MessageHandler, ConversationHandler, filters
from telegram.ext import MessageHandler, filters
from telegram.ext import MessageHandler, filters
import openpyxl
import os
import os
from telegram.ext import ConversationHandler
from fpdf import FPDF
from datetime import datetime
from io import BytesIO
import arabic_reshaper
from bidi.algorithm import get_display
import json
import pandas as pd
import os
PROJECT_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))  # یک پوشه بالاتر از فایل فعلی
#دخیره اطلاعات کاربر
PROFILE_FILE = os.path.join(os.path.dirname(__file__), "data", "profiles.json")

def load_profiles():
    if os.path.exists(PROFILE_FILE):
        with open(PROFILE_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}

def save_profiles(profiles):
    os.makedirs('data', exist_ok=True)
    with open(PROFILE_FILE, "w", encoding="utf-8") as f:
        json.dump(profiles, f, ensure_ascii=False, indent=2)

#متغیر رنگ و تعداد
WAITING_QUANTITY, WAITING_COLOR = range(2)
WAITING_EDIT_QUANTITY, WAITING_EDIT_COLOR = range(2,4)

# 🔐 توکن ربات
TOKEN = ""

# پیام خوش‌آمد هنگام شروع
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    known_users = context.application.bot_data.get("known_users", [])

    if not any(u["id"] == user.id for u in known_users):
        known_users.append({"id": user.id, "username": user.username})
        context.application.bot_data["known_users"] = known_users
    await update.message.reply_text("*سلام 😊 به ربات سفارش خوش اومدی!*\n\nبرای مشاهده منو روی این عبارت */menu* کیک کنید!\n\nبرای پیگیری خرید کلمه *پیگیری* رو تایپ کنید.\nبرای دریافت قیمت کالا کلمه *قیمت* رو تایپ کنید.\nبرای اطلاع از زمان و هزینه ارسال کلمه *ارسال* رو تایپ کنید.\nبرای دریافت موجودی کالا کلمه *موجودی* رو تایپ کنید.\nبرای سفارش و خرید کلمه *سفارش* رو تایپ کنید.",parse_mode="Markdown")

# منوی اصلی ربات
async def menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    keyboard = [
        ["🛍 مشاهده کاتالوگ"],
        ["📄 دریافت پیش‌فاکتور", "📞 تماس با پشتیبانی"],
        ["⚙️ تنظیمات حساب کاربری", "🛒 مشاهده سبد خرید"]
    ]
    reply_markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)
    await update.message.reply_text("لطفاً یک گزینه رو انتخاب کن:", reply_markup=reply_markup)

async def show_catalog_categories(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.message.text if update.message else update.callback_query.data

    # فقط اگر پیام "کاتالوگ" بود
    if query == "🛍 مشاهده کاتالوگ":
        buttons = [
            [InlineKeyboardButton("👗 مجلسی", callback_data="cat_w")],
            [InlineKeyboardButton("👖 شلوار شلوارک", callback_data="cat_p")],
            [InlineKeyboardButton("👕 ست اسپرت", callback_data="cat_s")],
            [InlineKeyboardButton("🧥 کت و هودی", callback_data="cat_h")],
            [InlineKeyboardButton("👕 تیشرت و پیراهن", callback_data="cat_t")],
            [InlineKeyboardButton("👜 اکسسوری و کیف", callback_data="cat_a")]
        ]
        markup = InlineKeyboardMarkup(buttons)

        await update.message.reply_text("لطفاً دسته‌بندی مورد نظر رو انتخاب کن:", reply_markup=markup)
async def category_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    category_map = {
        "cat_w": "🧵 محصولات مجلسی:",
        "cat_p": "👖 انواع شلوار و شلوارک:",
        "cat_s": "👕 ست‌های اسپرت:",
        "cat_h": "🧥 کت و هودی:",
        "cat_t":"👕 تیشرت و پیراهن",
        "cat_a": "🎒 اکسسوری و کیف:"
    }

    text = category_map.get(query.data, "دسته‌بندی یافت نشد.")
    await query.edit_message_text(text)

faq_answers = {
    "قیمت": "برای دریافت قیمت لطفاً از منوی 'مشاهده کاتالوگ‌ها' استفاده کن.",
    "ارسال": "ارسال سفارش‌ها معمولاً بین 5 تا 10 روز کاری زمان می‌بره ومعمولا بین 8 تا 12 درصد بسته به فاصله شهرتون هزینه ارسال داره.",
    "میرسه": "ارسال به سراسر کشور بین 5 تا 10 روز بسته به فاصله شهرتون انجام میشه.",
    "سفارش": "برای ثبت سفارش لطفاً از منوی /menu گزینه 'مشاهده کاتالوگ' رو انتخاب کن.",
    "موجودی": "موجودی اجناس هر روز به‌روز میشه. لطفاً از بخش کاتالوگ چک کن هرمدلی براتون نمایش داده میشه یعنی موجود هست.",
    "پیگیری": "چنانچه سفارشتون رو ثبت کردید میتونید با تماس پشتیبانی پیگیری کنید."
}

# تابع پاسخ‌دهی هوشمند
async def smart_reply(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_message = update.message.text.lower()
    response = None

    # جستجوی کلیدواژه
    for keyword in faq_answers:
        if keyword in user_message:
            response = faq_answers[keyword]
            break

    if response:
        await update.message.reply_text(response)
    else:
        await update.message.reply_text("متوجه نشدم 🧐 لطفاً دقیق‌تر بگو یا از منو استفاده کن.")
    
#تابع کاتالوگ ها 
async def show_category_products(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    category = query.data.replace("cat_", "")  # مثلا "cat_majlesi" → "majlesi"

    try:
        wb = openpyxl.load_workbook("catalog.xlsx")
        sheet = wb.active

        found = False
        for row in sheet.iter_rows(min_row=2, values_only=True):
            code, title, price, desc, color, image_file, cat, *_ = row

            if cat.strip() != category:
                continue  # فقط دسته مورد نظر رو نشون بده

            found = True
            caption = f"🧥 <b>{title}</b>\n" \
                      f"📦 کد کالا: <code>{code}</code>\n" \
                      f"💵 قیمت: <b>{price:,}</b> تومان\n" \
                      f"🎨 رنگبندی: <b>{color}</b>\n" \
                      f"📝 توضیحات: {desc}"

            full_path = os.path.join("images", image_file)
            keyboard = InlineKeyboardMarkup([[InlineKeyboardButton("➕ اضافه به سبد خرید", callback_data=f"add_{code}")]])

            with open(full_path, 'rb') as photo:
                await context.bot.send_photo(
                    chat_id=query.message.chat.id,
                    photo=photo,
                    caption=caption,
                    parse_mode="HTML",
                    reply_markup=keyboard
                )

        wb.close()
        if not found:
            await query.message.reply_text("❗ محصولی در این دسته‌بندی پیدا نشد.")
    
    except Exception as e:
        await query.message.reply_text(f"خطا در نمایش کاتالوگ دسته‌بندی:\n{str(e)}")

#تابع اضافه به سبد
async def add_to_cart(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    product_code = query.data.split("_")[1]
    context.user_data['current_product'] = product_code

    # از فایل اکسل تصویر محصول رو پیدا می‌کنیم
    wb = openpyxl.load_workbook("catalog.xlsx")
    sheet = wb.active

    image_file = None
    for row in sheet.iter_rows(min_row=2, values_only=True):
        code, title, price, desc, color, image, cat, *_ = row
        if str(code) == product_code:
            image_file = image
            break
    wb.close()

    context.user_data['current_image'] = image_file

    await query.message.reply_text(
        f"📦 چند عدد از محصول {product_code} می‌خوای؟\n\n"
        f"*مشتری عزیز؛\n*"
        f"*حداقل سفارش یک سری (یعنی 5 عدد)میباشد\n*"
        f"*از واردکردن تعداد کمتر از 5 عدد خودداری کنید\n*"
        f"*5عدد=یک بسته\n*"
        f"*10عدد=دو بسته\n*"
        f"و..."
        )
    return WAITING_QUANTITY

async def get_quantity(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        quantity = int(update.message.text)
        context.user_data['current_quantity'] = quantity
        await update.message.reply_text("🎨 حالا لطفاً رنگ مورد نظر رو وارد کن:")
        return WAITING_COLOR
    except ValueError:
        await update.message.reply_text("❌ لطفاً فقط عدد وارد کن:")
        return WAITING_QUANTITY
async def get_color(update: Update, context: ContextTypes.DEFAULT_TYPE):
    color = update.message.text

    product = context.user_data.get('current_product')
    quantity = context.user_data.get('current_quantity')
    image = context.user_data.get('current_image')

    # ساخت سبد خرید یا گرفتن قبلی
    cart = context.user_data.get("cart", [])
    cart.append({
        "product_code": product,
        "quantity": quantity,
        "image" : image,
        "color": color
    })
    context.user_data['cart'] = cart

    await update.message.reply_text(f"✅ {quantity} عدد از محصول {product} با رنگ {color} به سبد خرید اضافه شد.")
    
    # پاکسازی موقت
    context.user_data.pop('current_product', None)
    context.user_data.pop('current_quantity', None)
    context.user_data.pop('current_image', None)

    return ConversationHandler.END

#مشاهده سبد خرید
async def show_cart(update: Update, context: ContextTypes.DEFAULT_TYPE):
    cart = context.user_data.get("cart", [])
    if not cart:
        await update.message.reply_text("🛒 سبد خرید شما خالیه.")
        return
    chat_id = update.effective_chat.id
    bot = context.bot

    
    for index, item in enumerate(cart, start=1):
        caption = f"🔢 <b>#{index}</b>\n" \
               f"📦 کد محصول: <code>{item['product_code']}</code>\n" \
               f"🔢 تعداد: <b>{item['quantity']}</b>\n" \
               f"🎨 رنگ: <b>{item['color']}</b>\n\n"
        keyboard = InlineKeyboardMarkup([
            [
                InlineKeyboardButton("✏️ ویرایش تعداد", callback_data=f"edit_qty_{index-1}"),
                InlineKeyboardButton("🎨 ویرایش رنگ", callback_data=f"edit_color_{index-1}"),
                InlineKeyboardButton("🗑 حذف", callback_data=f"delete_{index-1}")
            ]
        ])
        try:
            image_path = os.path.join("images", item['image'])
            with open(image_path, 'rb') as photo:
                     await bot.send_photo(chat_id=chat_id, photo=photo, caption=caption, parse_mode="HTML", reply_markup=keyboard)
        except:
            await bot.send_message(chat_id=chat_id, text=caption, parse_mode="HTML", reply_markup=keyboard)
async def cart_callback_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    data = query.data
    cart = context.user_data.get("cart", [])

    if data.startswith("delete_"):
        idx = int(data.split("_")[1])
        if 0 <= idx < len(cart):
            deleted_item = cart.pop(idx)
            context.user_data["cart"] = cart
            await query.message.edit_caption(f"🗑 آیتم کد {deleted_item['product_code']} حذف شد.")
            # بعد نمایش مجدد سبد:
            await show_cart(update, context)

    elif data.startswith("edit_qty_"):
        idx = int(data.split("_")[2])
        context.user_data['edit_index'] = idx
        await query.message.reply_text("❓ تعداد جدید را وارد کنید:")
        return WAITING_EDIT_QUANTITY

    elif data.startswith("edit_color_"):
        idx = int(data.split("_")[2])
        context.user_data['edit_index'] = idx
        await query.message.reply_text("🎨 رنگ جدید را وارد کنید:")
        return WAITING_EDIT_COLOR

    return ConversationHandler.END
async def edit_quantity(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        new_qty = int(update.message.text)
        idx = context.user_data.get('edit_index')
        if idx is not None:
            cart = context.user_data.get("cart", [])
            if 0 <= idx < len(cart):
                cart[idx]['quantity'] = new_qty
                context.user_data["cart"] = cart
                await update.message.reply_text(f"✅ تعداد آیتم #{idx+1} به {new_qty} تغییر کرد.")
                context.user_data.pop('edit_index', None)
                # نمایش مجدد سبد
                await show_cart(update, context)
                return ConversationHandler.END
    except ValueError:
        await update.message.reply_text("❌ لطفاً عدد صحیح وارد کنید.")
        return WAITING_EDIT_QUANTITY

async def edit_color(update: Update, context: ContextTypes.DEFAULT_TYPE):
    new_color = update.message.text
    idx = context.user_data.get('edit_index')
    if idx is not None:
        cart = context.user_data.get("cart", [])
        if 0 <= idx < len(cart):
            cart[idx]['color'] = new_color
            context.user_data["cart"] = cart
            await update.message.reply_text(f"✅ رنگ آیتم #{idx+1} به '{new_color}' تغییر کرد.")
            context.user_data.pop('edit_index', None)
            await show_cart(update, context)
            return ConversationHandler.END

#هندلر سبد خرید کامل
conv_handler = ConversationHandler(
    entry_points=[
        CallbackQueryHandler(add_to_cart, pattern="^add_"),
        CallbackQueryHandler(cart_callback_handler, pattern="^(edit_qty_|edit_color_|delete_)")
    ],
    states={
        WAITING_QUANTITY: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_quantity)],
        WAITING_COLOR: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_color)],
        WAITING_EDIT_QUANTITY: [MessageHandler(filters.TEXT & ~filters.COMMAND, edit_quantity)],
        WAITING_EDIT_COLOR: [MessageHandler(filters.TEXT & ~filters.COMMAND, edit_color)],
    },
    fallbacks=[],)

# صدور پیش فاکتور
def generate_invoice(cart, username, profile):
    pdf = FPDF()
    pdf.add_page()
    pdf.add_font('BNazanin', '', 'font/BNazanin.ttf', uni=True)
    pdf.add_font('BNaznnBd', 'B', 'font/BNaznnBd.ttf', uni=True)
    pdf.add_font('Arial', '', 'font/arial.ttf', uni=True)
    pdf.set_font("BNazanin", size=12)

    # خواندن قیمت‌ها از فایل Excel
    df = pd.read_excel("catalog.xlsx")  # فرض بر این که ستون‌های code و price دارد
    price_dict = dict(zip(df['code'].astype(str), df['price']))

    total_price = 0

    # هدر
    pdf.set_font('BNaznnBd', 'B', 16)
    pdf.cell(200, 10, reshape_text("پیش‌فاکتور فروشگاه آنلاین ست"), ln=True, align="C")
    pdf.set_font('BNazanin', size=12)
    pdf.cell(200, 10, reshape_text(f"تاریخ: {datetime.now().strftime('%Y-%m-%d %H:%M')}"), ln=True, align="C")

    pdf.set_font('BNazanin', '', 12)
    pdf.cell(200, 10, reshape_text("نام کاربر:"), ln=False, align='R')
    pdf.set_font('Arial', '', 12)
    pdf.cell(0, 10, f"@{username}", ln=True, align='L')

    pdf.set_font('BNazanin', '', 12)
    pdf.cell(200, 10, reshape_text(f"نام: {profile['name']} -- شهر: {profile['city']}"), ln=True, align='R')
    pdf.cell(200, 10, reshape_text(f"تلفن: {profile['phone']}"), ln=True, align='R')
    pdf.multi_cell(0, 10, reshape_text(f"آدرس: {profile['address']}"), align='R')

    pdf.ln(10)

    # جدول
    pdf.set_font('BNaznnBd', 'B', 12)
    pdf.cell(30, 10, reshape_text("کد کالا"), border=1)
    pdf.cell(30, 10, reshape_text("رنگ"), border=1)
    pdf.cell(20, 10, reshape_text("تعداد"), border=1)
    pdf.cell(40, 10, reshape_text("قیمت واحد"), border=1)
    pdf.cell(40, 10, reshape_text("جمع ردیف"), border=1)
    pdf.ln()

    pdf.set_font("BNazanin", size=12)
    for item in cart:
        code = str(item["product_code"])
        color = item["color"]
        qty = item["quantity"]
        price = price_dict.get(code, 0)
        row_total = qty * price
        total_price += row_total

        pdf.cell(30, 10, code, border=1)
        pdf.cell(30, 10, reshape_text(color), border=1)
        pdf.cell(20, 10, str(qty), border=1)
        pdf.cell(40, 10, reshape_text(f"{price:,} تومان"), border=1)
        pdf.cell(40, 10, reshape_text(f"{row_total:,} تومان"), border=1)
        pdf.ln()

    # جمع کل نهایی
    pdf.set_font('BNaznnBd', 'B', 12)
    pdf.cell(150, 10, reshape_text("جمع کل"), border=1, align="R")
    pdf.cell(40, 10, reshape_text(f"{total_price:,} تومان"), border=1, align="C")

    pdf.ln(15)
    pdf.set_font("BNazanin", '', 10)
    pdf.multi_cell(0, 10, reshape_text("این پیش‌فاکتور به صورت سیستمی تولید شده است و نیاز به تایید نهایی فروشنده دارد.\nبا تشکر از خرید شما"))

    # ذخیره PDF
    safe_username = username.replace('@', '').replace(' ', '_')
    filename = f"invoice_{safe_username}.pdf"
    folder = os.path.join("tmp","")
    os.makedirs(folder, exist_ok=True)
    filepath = os.path.join(folder, filename)
    pdf.output(filepath)

    return filepath


# ارسال پیش فاکتور به مشتری و مدیر
admin_id = 199489709  # آیدی عددی ادمین

async def send_invoice(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = str(update.effective_user.id)
    profiles = context.application.bot_data.get("profiles", {})
    profile = profiles.get(user_id)

    if not profile:
        await update.message.reply_text("⚠️ لطفاً ابتدا اطلاعات حساب کاربری خود را کامل کنید.\nروی گزینه «⚙️ تنظیمات حساب کاربری» در منو کلیک کنید.")
        return

    if not profile:
        await update.message.reply_text("⚠️ لطفاً ابتدا اطلاعات حساب کاربری خود را کامل کنید.\nروی گزینه «⚙️ تنظیمات حساب کاربری» در منو کلیک کنید.")
        return

    cart = context.user_data.get("cart", [])
    if not cart:
        await update.message.reply_text("سبد خرید شما خالی است ❌")
        return

    user = update.effective_user
    username = update.effective_user.username or f"User{update.effective_user.id}"
    filepath = generate_invoice(cart, username, profile)

    with open(filepath, "rb") as f:
        pdf_bytes = BytesIO(f.read())
        pdf_bytes.seek(0)

        # 1️⃣ ارسال برای مشتری
        await context.bot.send_document(
            chat_id=update.effective_chat.id,
            document=pdf_bytes,
            filename=os.path.basename(filepath),
            caption="🧾 پیش‌فاکتور شما آماده است!"
        )

        # 2️⃣ ارسال برای ادمین
        pdf_bytes.seek(0)
        await context.bot.send_document(
            chat_id=admin_id,
            document=pdf_bytes,
            filename=os.path.basename(filepath),
            caption=f"📩 پیش‌فاکتور از کاربر: @{username}"
        )

    os.remove(filepath)
#تابع فارسی نویس
def reshape_text(text):
    reshaped_text = arabic_reshaper.reshape(text)
    bidi_text = get_display(reshaped_text)
    return bidi_text

# اجرای ربات
app = ApplicationBuilder().token(TOKEN).build()
app.bot_data["profiles"] = load_profiles()
app.add_handler(CommandHandler("start", start))
app.add_handler(CommandHandler("menu", menu))
app.add_handler(MessageHandler(filters.Text("🛍 مشاهده کاتالوگ"), show_catalog_categories))
app.add_handler(CallbackQueryHandler(show_category_products, pattern=r"^cat_"))
app.add_handler(conv_handler)
app.add_handler(MessageHandler(filters.TEXT & filters.Regex("مشاهده سبد خرید"), show_cart))
app.add_handler(MessageHandler(filters.TEXT & filters.Regex("دریافت پیش‌فاکتور"), send_invoice))

#اطلاعات کاربری
WAITING_NAME, WAITING_CITY, WAITING_PHONE, WAITING_ADDRESS = range(10, 14)

async def start_profile(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = str(update.effective_user.id)
    profiles = context.application.bot_data.get("profiles", {})
    profile = profiles.get(user_id)

    if profile:
        msg = f"👤 اطلاعات فعلی شما:\n\n"
        msg += f"🧍‍♂️ نام: {profile.get('name')}\n"
        msg += f"🏙 شهر: {profile.get('city')}\n"
        msg += f"📞 تلفن: {profile.get('phone')}\n"
        msg += f"📍 آدرس: {profile.get('address')}\n\n"
        msg += "❓آیا مایل به ویرایش اطلاعات هستید؟ (بله / خیر)"
        await update.message.reply_text(msg)
        return "CONFIRM_EDIT"

    else:
        await update.message.reply_text("📝 لطفاً نام کامل خود را وارد کنید:")
        return WAITING_NAME


async def get_name(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["profile"] = {"name": update.message.text}
    await update.message.reply_text("🏙 شهر خود را وارد کنید:")
    return WAITING_CITY

async def get_city(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["profile"]["city"] = update.message.text
    await update.message.reply_text("📞 شماره تلفن خود را وارد کنید:")
    return WAITING_PHONE

async def get_phone(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["profile"]["phone"] = update.message.text
    await update.message.reply_text("📍 آدرس کامل پستی خود را وارد کنید:")
    return WAITING_ADDRESS

async def get_address(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = str(update.effective_user.id)
    address = update.message.text

    profile = context.user_data.get("profile", {})
    profile["address"] = address
    context.user_data["profile"] = profile

    # ذخیره در فایل JSON
    profiles = context.application.bot_data.get("profiles", {})
    profiles[user_id] = profile
    context.application.bot_data["profiles"] = profiles
    save_profiles(profiles)
    
    await update.message.reply_text("✅ اطلاعات حساب شما با موفقیت ذخیره شد.")
    return ConversationHandler.END

#تابع ویرایش اطلاعات کاربری
async def confirm_edit(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip().lower()
    if text in ["بله", "بلی", "آره", "yes"]:
        await update.message.reply_text("📝 لطفاً نام کامل خود را وارد کنید:")
        return WAITING_NAME
    else:
        await update.message.reply_text("👌 اطلاعات فعلی شما حفظ خواهد شد.")
        return ConversationHandler.END


profile_handler = ConversationHandler(
    entry_points=[MessageHandler(filters.TEXT & filters.Regex("⚙️ تنظیمات حساب کاربری"), start_profile)],
    states={
        "CONFIRM_EDIT": [MessageHandler(filters.TEXT & ~filters.COMMAND, confirm_edit)],
        WAITING_NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_name)],
        WAITING_CITY: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_city)],
        WAITING_PHONE: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_phone)],
        WAITING_ADDRESS: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_address)],
    },
    fallbacks=[],
)
app.add_handler(profile_handler)
app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, smart_reply))

print("bot is running... ⏳")
app.run_polling()